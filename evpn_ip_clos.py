#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EVPN IP CLOS Fabric Builder - Version 0.2
=== USE AT YOUR OWN RISK! ===
Features:
- Reads Excel (FABRIC, SWITCHES, NETWORKS, WAN)
- Pod-aware with multi-pod support
-- Flexible pod topologies: 3-tier (Access->Distribution->Core) or 2-tier (Access->Core)
-- Each pod can independently use either topology in multi-pod deployments
- Optic port configuration (speed and channelization)
- IPv4 or IPv6 underlay
- Dual-Stack in the Overlay
- DHCP relay support for IPv4 and IPv6
- WAN L3 routing on border/core switches with BGP peering if WAN is filled out
-- Routing policy explicit deny support (EXPORT_EXPLICT_DENY and IMPORT_EXPLICT_DENY flags)
-- Auto-generated export policy terms match on prefix only (no protocol matching)
- Detects existing topologies by name and UPDATES them instead of creating new ones

Author: Lukas Eisenberger (leisenberger@juniper.net)

"""

import sys
import ipaddress
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook

import mistClient
import mistHelpers


# =============================================================================
# CONFIG
# =============================================================================
MIST_ORGID = "ORG_ID"
MIST_TOKEN = "API-TOKEN"
MIST_API_URL = "https://api.eu.mist.com"
spreadsheetname = "evpn_ip_clos.xlsx"
# =============================================================================


def _validate_required(value: Any, field_name: str, context: str = "") -> Any:
    """Validate that a required field is present and not empty"""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        ctx = f" in {context}" if context else ""
        raise ValueError(f"Required field '{field_name}'{ctx} is missing or empty")
    return value


def _b(v, field_name: str, context: str = ""):
    """Convert to boolean - STRICT: raises exception if not parseable"""
    if isinstance(v, bool):
        return v
    if v is None or (isinstance(v, str) and v.strip() == ""):
        raise ValueError(f"Boolean field '{field_name}' in {context} is missing or empty")

    s = str(v).strip().lower()
    if s in ("1", "true", "t", "y", "yes", "on"):
        return True
    elif s in ("0", "false", "f", "n", "no", "off"):
        return False
    else:
        raise ValueError(f"Cannot parse '{v}' as boolean for field '{field_name}' in {context}")


def _b_optional(v, field_name: str, context: str = "", default: bool = False) -> bool:
    """Convert to boolean - OPTIONAL: returns default if empty, raises exception if not parseable"""
    if isinstance(v, bool):
        return v
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default

    s = str(v).strip().lower()
    if s in ("1", "true", "t", "y", "yes", "on"):
        return True
    elif s in ("0", "false", "f", "n", "no", "off"):
        return False
    else:
        raise ValueError(f"Cannot parse '{v}' as boolean for field '{field_name}' in {context}")


def _i(v, field_name: str, context: str = ""):
    """Convert to integer - STRICT: raises exception if not parseable"""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        raise ValueError(f"Integer field '{field_name}' in {context} is missing or empty")

    try:
        if isinstance(v, int):
            return v
        if isinstance(v, float):
            return int(v)
        return int(str(v).strip())
    except (ValueError, TypeError) as e:
        raise ValueError(f"Cannot parse '{v}' as integer for field '{field_name}' in {context}: {e}")


def _parse_ip_list(ip_str: Optional[str], field_name: str, af: int, context: str = "") -> List[str]:
    """
    Parse comma or space-separated IP addresses
    af: 4 for IPv4, 6 for IPv6
    Returns list of validated IP addresses
    Raises exception if duplicate IPs are found
    """
    if not ip_str or (isinstance(ip_str, str) and ip_str.strip() == ""):
        return []

    result = []
    seen = set()
    ip_str = str(ip_str).strip()

    # Split by comma or whitespace
    ips = [ip.strip() for ip in ip_str.replace(',', ' ').split() if ip.strip()]

    for ip in ips:
        try:
            addr = ipaddress.ip_address(ip)
            if (af == 4 and addr.version != 4) or (af == 6 and addr.version != 6):
                raise ValueError(f"Expected IPv{af} address but got IPv{addr.version}")

            # Use normalized IP address string for duplicate detection
            normalized_ip = str(addr)

            # Check for duplicates and raise exception
            if normalized_ip in seen:
                ctx = f" in {context}" if context else ""
                raise ValueError(
                    f"Duplicate IPv{af} address '{normalized_ip}' found in field '{field_name}'{ctx}. "
                    f"Each IP address must be listed only once."
                )

            result.append(normalized_ip)
            seen.add(normalized_ip)
        except ValueError as e:
            ctx = f" in {context}" if context else ""
            if "Duplicate" in str(e):
                raise  # Re-raise duplicate error as-is
            raise ValueError(f"Invalid IPv{af} address '{ip}' in field '{field_name}'{ctx}: {e}")

    return result


def _clean_excel_string(s: str) -> str:
    """
    Clean Excel string by removing carriage returns and other special characters.
    Excel often includes \r\n or \r (CR/LF) which appear as _x000D_ in XML.
    """
    if not s:
        return s
    # Remove carriage returns and normalize line endings
    s = s.replace('\r\n', '\n').replace('\r', '\n')
    # Remove any remaining special characters
    s = s.replace('_x000D_', '')
    return s


def _parse_static_routes(routes_str: Optional[str], field_name: str) -> Dict[str, Dict[str, str]]:
    """
    Parse static routes in format: route@nexthop route@nexthop ...
    Examples:
      "0.0.0.0/0@192.168.10.254"
      "192.168.66.0/24@192.168.10.254 192.168.77.0/24@192.168.10.254"
      "::/0@2001:10::254"
    Returns dict: {route: {"via": nexthop}, ...}
    Note: Duplicate routes will use the last nexthop specified

    STRICT VALIDATION: All routes must be valid or the entire field is rejected
    """
    result: Dict[str, Dict[str, str]] = {}
    if not routes_str or (isinstance(routes_str, str) and routes_str.strip() == ""):
        return result

    routes_str = str(routes_str).strip()
    pairs = routes_str.split()

    for pair in pairs:
        pair = pair.strip()
        if not pair:
            continue

        if '@' not in pair:
            raise ValueError(
                f"Error parsing static routes in field '{field_name}': "
                f"Invalid format '{pair}' - must be 'route@nexthop'. "
                f"Example: '0.0.0.0/0@192.168.10.1' or '192.168.1.0/24@10.0.0.1'"
            )

        parts = pair.split('@', 1)
        if len(parts) != 2:
            raise ValueError(
                f"Error parsing static routes in field '{field_name}': "
                f"Invalid format '{pair}' - must be 'route@nexthop'"
            )

        route = parts[0].strip()
        nexthop = parts[1].strip()

        if not route:
            raise ValueError(
                f"Error parsing static routes in field '{field_name}': "
                f"Empty route in '{pair}'"
            )
        if not nexthop:
            raise ValueError(
                f"Error parsing static routes in field '{field_name}': "
                f"Empty nexthop in '{pair}'"
            )

        # Validate route is a valid network with prefix
        try:
            route_network = ipaddress.ip_network(route, strict=False)
        except ValueError as e:
            raise ValueError(
                f"Error parsing static routes in field '{field_name}': "
                f"Invalid network '{route}' in route '{pair}'. "
                f"Network must be in CIDR format (e.g., '192.168.1.0/24' or '10.0.0.0/8'). "
                f"Details: {e}"
            )

        # Validate nexthop is a valid IP address
        try:
            nexthop_ip = ipaddress.ip_address(nexthop)
        except ValueError as e:
            raise ValueError(
                f"Error parsing static routes in field '{field_name}': "
                f"Invalid nexthop IP '{nexthop}' in route '{pair}'. "
                f"Nexthop must be a valid IP address (e.g., '192.168.1.1' or '2001::1'). "
                f"Details: {e}"
            )

        # Validate route and nexthop are same IP version
        if route_network.version != nexthop_ip.version:
            raise ValueError(
                f"Error parsing static routes in field '{field_name}': "
                f"IP version mismatch in route '{pair}': "
                f"route '{route}' is IPv{route_network.version} but "
                f"nexthop '{nexthop}' is IPv{nexthop_ip.version}. "
                f"Both must be the same IP version."
            )

        # Use the normalized CIDR notation for the route
        normalized_route = str(route_network)
        result[normalized_route] = {"via": str(nexthop_ip)}

    return result


def _parse_optic_config(optic_str: Optional[str]) -> Dict[str, Dict[str, Any]]:
    """
    Parse optic port config in format: port=speed=channelized port=speed=channelized ...
    Examples:
      "et-0/0/5=25g=true et-0/0/12=10g=true"
      "et-0/0/16=100g=false"
    Returns dict: {port: {"speed": speed, "channelized": bool}, ...}
    """
    result: Dict[str, Dict[str, Any]] = {}
    if not optic_str or (isinstance(optic_str, str) and optic_str.strip() == ""):
        return result

    try:
        optic_str = str(optic_str).strip()
        entries = optic_str.split()

        for entry in entries:
            entry = entry.strip()
            if not entry:
                continue

            parts = entry.split('=')
            if len(parts) != 3:
                raise ValueError(f"Invalid format '{entry}' - must be 'port=speed=channelized'")

            port = parts[0].strip()
            speed = parts[1].strip().lower()
            channelized_str = parts[2].strip().lower()

            if not port:
                raise ValueError(f"Empty port in '{entry}'")
            if not speed:
                raise ValueError(f"Empty speed in '{entry}'")

            # Validate speed
            valid_speeds = ["10g", "25g", "50g", "100g", "200g", "auto"]
            if speed not in valid_speeds:
                raise ValueError(f"Invalid speed '{speed}' - must be one of {valid_speeds}")

            # Parse channelized
            if channelized_str in ("true", "t", "1", "yes", "y"):
                channelized = True
            elif channelized_str in ("false", "f", "0", "no", "n"):
                channelized = False
            else:
                raise ValueError(f"Invalid channelized value '{channelized_str}' - must be true/false")

            result[port] = {"speed": speed, "channelized": channelized}

    except Exception as e:
        raise ValueError(f"Error parsing optic config: {e}")

    return result


def _parse_fabric(wb) -> Dict[str, Any]:
    """Parse FABRIC sheet for global configuration - STRICT validation"""
    if "FABRIC" not in wb.sheetnames:
        raise Exception("FABRIC sheet missing in workbook")

    sh = wb["FABRIC"]
    data = {}

    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        val = row[1]
        data[key] = val

    try:
        topology_name = _validate_required(
            data.get("topology_name") or data.get("topologyname"),
            "topology_name",
            "FABRIC sheet"
        )
        site_name = _validate_required(
            data.get("site_name") or data.get("sitename") or data.get("site"),
            "site_name",
            "FABRIC sheet"
        )

        bgp_as_pool_range_for_underlay = _i(
            data.get("bgp_as_pool_range_for_underlay"),
            "bgp_as_pool_range_for_underlay",
            "FABRIC sheet"
        )
        use_ipv6_underlay = _b(data.get("use_ipv6_underlay"), "use_ipv6_underlay", "FABRIC sheet")
        core_as_border = _b(data.get("core_as_border"), "core_as_border", "FABRIC sheet")

        underlay_subnet = None
        underlay_subnet_raw = data.get("underlay_subnet")

        if use_ipv6_underlay:
            # IPv6 underlay REQUIRES underlay_subnet
            underlay_subnet = _validate_required(
                underlay_subnet_raw,
                "underlay_subnet",
                "FABRIC sheet (required when use_ipv6_underlay=true)"
            )
            try:
                ipaddress.ip_network(underlay_subnet)
            except ValueError as e:
                raise ValueError(f"Invalid underlay_subnet '{underlay_subnet}': {e}")
        elif underlay_subnet_raw:
            # IPv4 underlay: underlay_subnet is optional but validate if present
            underlay_subnet = str(underlay_subnet_raw).strip()
            try:
                ipaddress.ip_network(underlay_subnet)
            except ValueError as e:
                raise ValueError(f"Invalid underlay_subnet '{underlay_subnet}': {e}")

        auto_router_id_subnet = data.get("auto_router_id_subnet")
        if auto_router_id_subnet:
            try:
                ipaddress.ip_network(str(auto_router_id_subnet).strip())
            except ValueError as e:
                raise ValueError(f"Invalid auto_router_id_subnet '{auto_router_id_subnet}': {e}")

        auto_router_id_subnet6 = data.get("auto_router_id_subnet6")
        if auto_router_id_subnet6:
            try:
                ipaddress.ip_network(str(auto_router_id_subnet6).strip())
            except ValueError as e:
                raise ValueError(f"Invalid auto_router_id_subnet6 '{auto_router_id_subnet6}': {e}")

        auto_loopback_subnet = data.get("auto_loopback_subnet")
        if auto_loopback_subnet:
            try:
                ipaddress.ip_network(str(auto_loopback_subnet).strip())
            except ValueError as e:
                raise ValueError(f"Invalid auto_loopback_subnet '{auto_loopback_subnet}': {e}")

        auto_loopback_subnet6 = data.get("auto_loopback_subnet6")
        if auto_loopback_subnet6:
            try:
                ipaddress.ip_network(str(auto_loopback_subnet6).strip())
            except ValueError as e:
                raise ValueError(f"Invalid auto_loopback_subnet6 '{auto_loopback_subnet6}': {e}")

        return {
            "topology_name": topology_name,
            "site_name": site_name,
            "bgp_as_pool_range_for_underlay": bgp_as_pool_range_for_underlay,
            "use_ipv6_underlay": use_ipv6_underlay,
            "core_as_border": core_as_border,
            "underlay_subnet": underlay_subnet,
            "auto_router_id_subnet": auto_router_id_subnet,
            "auto_router_id_subnet6": auto_router_id_subnet6,
            "auto_loopback_subnet": auto_loopback_subnet,
            "auto_loopback_subnet6": auto_loopback_subnet6,
        }

    except Exception as e:
        raise Exception(f"Error parsing FABRIC sheet: {e}")


def _parse_switches(wb) -> List[Dict[str, Any]]:
    """
    Parse SWITCHES sheet - STRICT validation
    Expected columns: HOSTNAME, ROLE, POD, UPLINKS, UPLINK_PORTS, DOWNLINKS, DOWNLINK_PORTS, OPTIC_CONFIG
    Supported roles: core, distribution, access, border

    Note: Row 1 contains comments, Row 2 contains headers, data starts at Row 3
    """
    if "SWITCHES" not in wb.sheetnames:
        raise Exception("SWITCHES sheet missing in workbook")

    sh = wb["SWITCHES"]
    # Row 1 is comments, Row 2 is headers
    header = [str(c).strip().upper() if c else "" for c in next(sh.iter_rows(min_row=2, max_row=2, values_only=True))]

    required_columns = ["HOSTNAME", "ROLE"]
    for col in required_columns:
        if col not in header:
            raise Exception(f"SWITCHES sheet missing required column: {col}")

    switches = []
    hostnames_seen = set()
    row_num = 3  # Data starts at row 3

    for row in sh.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            row_num += 1
            continue

        try:
            row_data = {header[i]: row[i] for i in range(min(len(header), len(row)))}

            hostname = _validate_required(row_data.get("HOSTNAME"), "HOSTNAME", f"SWITCHES row {row_num}")
            hostname = str(hostname).strip()

            # Check for duplicate hostname
            if hostname in hostnames_seen:
                raise ValueError(f"Duplicate HOSTNAME '{hostname}' found in SWITCHES row {row_num}")
            hostnames_seen.add(hostname)

            role = _validate_required(row_data.get("ROLE"), "ROLE", f"SWITCHES row {row_num}")
            role = str(role).strip().lower()

            if role not in ["core", "distribution", "access", "border"]:
                raise ValueError(
                    f"Invalid ROLE '{role}' in SWITCHES row {row_num} - must be 'core', 'distribution', 'access', or 'border'"
                )

            # Pod is optional but if present must be valid integer
            # Note: Border switches should NOT have a pod assigned
            pod = None
            pod_raw = row_data.get("POD")
            if pod_raw is not None and str(pod_raw).strip() != "":
                pod = _i(pod_raw, "POD", f"SWITCHES row {row_num}")
                if role == "border":
                    raise ValueError(f"Border switch '{hostname}' (row {row_num}) should not have a POD assigned")

            uplinks_str = str(row_data.get("UPLINKS") or "").strip()
            uplinks = [u.strip() for u in uplinks_str.split(",") if u.strip()] if uplinks_str else []

            uplink_ports_str = str(row_data.get("UPLINK_PORTS") or "").strip()
            uplink_ports = uplink_ports_str if uplink_ports_str else ""

            downlinks_str = str(row_data.get("DOWNLINKS") or "").strip()
            downlinks = [d.strip() for d in downlinks_str.split(",") if d.strip()] if downlinks_str else []

            downlink_ports_str = str(row_data.get("DOWNLINK_PORTS") or "").strip()
            downlink_ports = downlink_ports_str if downlink_ports_str else ""

            # Parse optic config
            optic_config_str = str(row_data.get("OPTIC_CONFIG") or "").strip()
            optic_config = _parse_optic_config(optic_config_str)

            # Validate role-specific requirements
            if role in ["distribution", "access"] and not uplinks:
                raise ValueError(f"Switch '{hostname}' with role '{role}' must have UPLINKS defined")

            if role in ["core", "distribution"] and not downlinks:
                raise ValueError(f"Switch '{hostname}' with role '{role}' must have DOWNLINKS defined")

            if role == "border":
                if uplinks:
                    raise ValueError(f"Border switch '{hostname}' (row {row_num}) should not have UPLINKS defined")
                if uplink_ports:
                    raise ValueError(f"Border switch '{hostname}' (row {row_num}) should not have UPLINK_PORTS defined")
                if not downlinks:
                    raise ValueError(
                        f"Border switch '{hostname}' (row {row_num}) must have DOWNLINKS defined (to core switches)"
                    )

            if uplinks and not uplink_ports:
                raise ValueError(f"Switch '{hostname}' has UPLINKS but no UPLINK_PORTS defined")

            if downlinks and not downlink_ports:
                raise ValueError(f"Switch '{hostname}' has DOWNLINKS but no DOWNLINK_PORTS defined")

            switches.append({
                "hostname": hostname,
                "role": role,
                "pod": pod,
                "uplinks": uplinks,
                "uplink_ports": uplink_ports,
                "downlinks": downlinks,
                "downlink_ports": downlink_ports,
                "optic_config": optic_config
            })

        except Exception as e:
            raise Exception(f"Error parsing SWITCHES row {row_num}: {e}")

        row_num += 1

    if not switches:
        raise Exception("SWITCHES sheet has no valid data rows")

    return switches


def _parse_networks(wb, use_ipv6_underlay: bool) -> Tuple[Dict[str, Any], Dict[str, Any], List[str], Dict[str, Any]]:
    """
    Parse NETWORKS sheet - STRICT validation with comprehensive syntax checking

    Expected columns: NETWORKNAME, VLAN_ID, VRF, GATEWAY, GATEWAY6, STATIC_ROUTESv4, STATIC_ROUTESv6, DHCP, DHCPv6

    Note: Row 1 contains comments, Row 2 contains headers, data starts at Row 3

    Returns: (site_networks, vrf_instances, network_name_list, dhcpd_config)
    """
    if "NETWORKS" not in wb.sheetnames:
        raise Exception("NETWORKS sheet missing in workbook")

    sh = wb["NETWORKS"]
    # Row 1 is comments, Row 2 is headers
    header = [str(c).strip().upper() if c else "" for c in next(sh.iter_rows(min_row=2, max_row=2, values_only=True))]

    required_columns = ["NETWORKNAME", "VLAN_ID"]
    for col in required_columns:
        if col not in header:
            raise Exception(f"NETWORKS sheet missing required column: {col}")

    site_networks: Dict[str, Any] = {}
    vrf_instances: Dict[str, Any] = {}
    network_name_list: List[str] = []
    dhcpd_config: Dict[str, Any] = {}
    vlan_ids_seen: Dict[int, str] = {}  # Track VLAN IDs and which network uses them

    row_num = 3  # Data starts at row 3
    for row in sh.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            row_num += 1
            continue

        try:
            net = {header[i]: row[i] for i in range(min(len(header), len(row)))}

            name = _validate_required(net.get("NETWORKNAME"), "NETWORKNAME", f"NETWORKS row {row_num}")
            name = str(name).strip()

            if name in network_name_list:
                raise ValueError(f"Duplicate NETWORKNAME '{name}' in NETWORKS row {row_num}")

            network_name_list.append(name)

            vlan_id = _i(net.get("VLAN_ID"), "VLAN_ID", f"NETWORKS row {row_num}")

            if vlan_id < 1 or vlan_id > 4094:
                raise ValueError(f"VLAN_ID {vlan_id} out of valid range (1-4094) in NETWORKS row {row_num}")

            # Check for duplicate VLAN IDs
            if vlan_id in vlan_ids_seen:
                raise ValueError(
                    f"Duplicate VLAN_ID {vlan_id} in NETWORKS row {row_num}: "
                    f"already used by network '{vlan_ids_seen[vlan_id]}'"
                )
            vlan_ids_seen[vlan_id] = name

            vrf = str(net.get("VRF") or "").strip()
            gw4 = net.get("GATEWAY")
            gw6 = net.get("GATEWAY6")

            static_routes_v4_str = net.get("STATIC_ROUTESv4") or net.get("STATIC_ROUTESV4")
            static_routes_v6_str = net.get("STATIC_ROUTESv6") or net.get("STATIC_ROUTESV6")

            dhcp_servers_str = net.get("DHCP")
            dhcpv6_servers_str = net.get("DHCPv6") or net.get("DHCPV6")

            # Check IPv6 configuration when underlay is disabled
            if not use_ipv6_underlay:
                if gw6 and str(gw6).strip():
                    raise ValueError(
                        f"Network '{name}' (row {row_num}): GATEWAY6 is configured but use_ipv6_underlay=false in FABRIC sheet. "
                        f"IPv6 gateways can only be configured when IPv6 underlay is enabled."
                    )

                if static_routes_v6_str and str(static_routes_v6_str).strip():
                    raise ValueError(
                        f"Network '{name}' (row {row_num}): STATIC_ROUTESv6 is configured but use_ipv6_underlay=false in FABRIC sheet. "
                        f"IPv6 static routes can only be configured when IPv6 underlay is enabled."
                    )

                if dhcpv6_servers_str and str(dhcpv6_servers_str).strip():
                    raise ValueError(
                        f"Network '{name}' (row {row_num}): DHCPv6 is configured but use_ipv6_underlay=false in FABRIC sheet. "
                        f"DHCPv6 relay can only be configured when IPv6 underlay is enabled."
                    )

            subnet4 = ""
            subnet6 = ""
            gw4_ip = ""
            gw6_ip = ""

            # Parse IPv4 gateway if present
            if gw4 and str(gw4).strip():
                try:
                    i4 = ipaddress.ip_interface(str(gw4).strip())
                    subnet4 = f"{i4.network.network_address}/{i4.network.prefixlen}"
                    gw4_ip = str(i4.ip)

                    # Validate gateway is within subnet
                    network4 = i4.network
                    gateway_addr4 = i4.ip

                    # Check if gateway is the network address (first address)
                    if gateway_addr4 == network4.network_address:
                        raise ValueError(
                            f"Invalid IPv4 GATEWAY '{gw4}' in NETWORKS row {row_num}: "
                            f"Gateway IP {gw4_ip} is the network address of {subnet4}. "
                            f"Gateway cannot be the network address (first IP in subnet)."
                        )

                    # Check if gateway is the broadcast address (last address)
                    if gateway_addr4 == network4.broadcast_address:
                        raise ValueError(
                            f"Invalid IPv4 GATEWAY '{gw4}' in NETWORKS row {row_num}: "
                            f"Gateway IP {gw4_ip} is the broadcast address of {subnet4}. "
                            f"Gateway cannot be the broadcast address (last IP in subnet)."
                        )

                except ValueError as e:
                    # Re-raise our custom errors or wrap parsing errors
                    if "network address" in str(e) or "broadcast address" in str(e):
                        raise
                    raise ValueError(f"Invalid IPv4 GATEWAY '{gw4}' in NETWORKS row {row_num}: {e}")

            # Parse IPv6 gateway if present
            if gw6 and str(gw6).strip():
                try:
                    i6 = ipaddress.ip_interface(str(gw6).strip())
                    subnet6 = f"{i6.network.network_address}/{i6.network.prefixlen}"
                    gw6_ip = str(i6.ip)

                    # Validate gateway is within subnet
                    network6 = i6.network
                    gateway_addr6 = i6.ip

                    # IPv6 doesn't have broadcast, but check for network address
                    if gateway_addr6 == network6.network_address:
                        raise ValueError(
                            f"Invalid IPv6 GATEWAY6 '{gw6}' in NETWORKS row {row_num}: "
                            f"Gateway IP {gw6_ip} is the network address of {subnet6}. "
                            f"Gateway cannot be the network address (first IP in subnet)."
                        )

                except ValueError as e:
                    # Re-raise our custom errors or wrap parsing errors
                    if "network address" in str(e):
                        raise
                    raise ValueError(f"Invalid IPv6 GATEWAY6 '{gw6}' in NETWORKS row {row_num}: {e}")

            # L2-only validation: comprehensive check for Layer 3 settings
            is_l2_only = not gw4_ip and not gw6_ip

            if is_l2_only:
                # L2-only networks should ONLY have NETWORKNAME and VLAN_ID
                l2_violations = []

                if vrf:
                    l2_violations.append(f"VRF '{vrf}'")

                if static_routes_v4_str and str(static_routes_v4_str).strip():
                    l2_violations.append(f"STATIC_ROUTESv4 '{static_routes_v4_str}'")

                if static_routes_v6_str and str(static_routes_v6_str).strip():
                    l2_violations.append(f"STATIC_ROUTESv6 '{static_routes_v6_str}'")

                if dhcp_servers_str and str(dhcp_servers_str).strip():
                    l2_violations.append(f"DHCP '{dhcp_servers_str}'")

                if dhcpv6_servers_str and str(dhcpv6_servers_str).strip():
                    l2_violations.append(f"DHCPv6 '{dhcpv6_servers_str}'")

                if l2_violations:
                    violations_text = ", ".join(l2_violations)
                    raise ValueError(
                        f"Network '{name}' (row {row_num}) is L2-only (no gateways configured) but has Layer 3 settings defined: {violations_text}. "
                        f"L2-only networks should ONLY have NETWORKNAME and VLAN_ID configured."
                    )

            if not is_l2_only and not vrf:
                raise ValueError(
                    f"Network '{name}' (row {row_num}) has gateways configured but no VRF assigned. "
                    f"Layer 3 networks must be assigned to a VRF."
                )

            # Build network definition
            site_networks[name] = {"vlan_id": vlan_id}

            if subnet4:
                site_networks[name]["subnet"] = subnet4
            if subnet6:
                site_networks[name]["subnet6"] = subnet6
            if gw4_ip:
                site_networks[name]["gateway"] = gw4_ip
            if gw6_ip:
                site_networks[name]["gateway6"] = gw6_ip

            # Handle VRF assignment and static routes
            if vrf:
                inst = vrf_instances.setdefault(vrf, {"networks": []})
                inst["networks"].append(name)

                # Parse static routes
                if static_routes_v4_str and str(static_routes_v4_str).strip():
                    routes_v4 = _parse_static_routes(static_routes_v4_str, f"STATIC_ROUTESv4 in row {row_num}")
                    if routes_v4:
                        inst.setdefault("extra_routes", {}).update(routes_v4)

                if static_routes_v6_str and str(static_routes_v6_str).strip():
                    routes_v6 = _parse_static_routes(static_routes_v6_str, f"STATIC_ROUTESv6 in row {row_num}")
                    if routes_v6:
                        inst.setdefault("extra_routes6", {}).update(routes_v6)

            # Handle DHCP relay configuration
            dhcp_v4_servers = []
            dhcp_v6_servers = []

            if dhcp_servers_str and str(dhcp_servers_str).strip():
                dhcp_v4_servers = _parse_ip_list(
                    dhcp_servers_str,
                    "DHCP",
                    4,
                    f"NETWORKS row {row_num}"
                )

            if dhcpv6_servers_str and str(dhcpv6_servers_str).strip():
                dhcp_v6_servers = _parse_ip_list(
                    dhcpv6_servers_str,
                    "DHCPv6",
                    6,
                    f"NETWORKS row {row_num}"
                )

            # Add to dhcpd_config if any relay servers are defined
            if dhcp_v4_servers or dhcp_v6_servers:
                dhcp_entry = {
                    "type": "relay"
                }

                if dhcp_v4_servers:
                    dhcp_entry["servers"] = dhcp_v4_servers

                if dhcp_v6_servers:
                    dhcp_entry["type6"] = "relay"
                    dhcp_entry["serversv6"] = dhcp_v6_servers

                dhcpd_config[name] = dhcp_entry

        except Exception as e:
            raise Exception(f"Error parsing NETWORKS row {row_num}: {e}")

        row_num += 1

    if not network_name_list:
        raise Exception("NETWORKS sheet has no valid data rows")

    # =========================================================================
    # STRICT VALIDATION: Check for conflicts within VRFs
    # =========================================================================

    # Track gateways and subnets per VRF for validation
    vrf_gateways_v4: Dict[str, Dict[str, str]] = {}  # {vrf: {gateway_ip: network_name}}
    vrf_gateways_v6: Dict[str, Dict[str, str]] = {}  # {vrf: {gateway_ip: network_name}}
    vrf_subnets_v4: Dict[str, List[Tuple[str, str]]] = {}  # {vrf: [(network_name, subnet)]}
    vrf_subnets_v6: Dict[str, List[Tuple[str, str]]] = {}  # {vrf: [(network_name, subnet)]}

    # Collect gateway and subnet data per VRF
    for net_name, net_data in site_networks.items():
        # Find which VRF this network belongs to
        net_vrf = None
        for vrf_name, vrf_data in vrf_instances.items():
            if net_name in vrf_data.get("networks", []):
                net_vrf = vrf_name
                break

        if not net_vrf:
            continue  # L2-only network, skip VRF checks

        # Check IPv4 gateway duplicates
        if net_data.get("gateway"):
            gw4 = net_data["gateway"]
            if net_vrf not in vrf_gateways_v4:
                vrf_gateways_v4[net_vrf] = {}

            if gw4 in vrf_gateways_v4[net_vrf]:
                raise ValueError(
                    f"Duplicate IPv4 gateway '{gw4}' in VRF '{net_vrf}': "
                    f"used by both network '{vrf_gateways_v4[net_vrf][gw4]}' and '{net_name}'. "
                    f"Each network in a VRF must have a unique gateway IP."
                )
            vrf_gateways_v4[net_vrf][gw4] = net_name

        # Check IPv6 gateway duplicates
        if net_data.get("gateway6"):
            gw6 = net_data["gateway6"]
            if net_vrf not in vrf_gateways_v6:
                vrf_gateways_v6[net_vrf] = {}

            if gw6 in vrf_gateways_v6[net_vrf]:
                raise ValueError(
                    f"Duplicate IPv6 gateway '{gw6}' in VRF '{net_vrf}': "
                    f"used by both network '{vrf_gateways_v6[net_vrf][gw6]}' and '{net_name}'. "
                    f"Each network in a VRF must have a unique gateway IP."
                )
            vrf_gateways_v6[net_vrf][gw6] = net_name

        # Collect IPv4 subnets for overlap checking
        if net_data.get("subnet"):
            if net_vrf not in vrf_subnets_v4:
                vrf_subnets_v4[net_vrf] = []
            vrf_subnets_v4[net_vrf].append((net_name, net_data["subnet"]))

        # Collect IPv6 subnets for overlap checking
        if net_data.get("subnet6"):
            if net_vrf not in vrf_subnets_v6:
                vrf_subnets_v6[net_vrf] = []
            vrf_subnets_v6[net_vrf].append((net_name, net_data["subnet6"]))

    # Check for IPv4 subnet overlaps within each VRF
    for vrf_name, subnets in vrf_subnets_v4.items():
        for i in range(len(subnets)):
            net_name1, subnet1_str = subnets[i]
            subnet1 = ipaddress.ip_network(subnet1_str, strict=False)

            for j in range(i + 1, len(subnets)):
                net_name2, subnet2_str = subnets[j]
                subnet2 = ipaddress.ip_network(subnet2_str, strict=False)

                if subnet1.overlaps(subnet2):
                    raise ValueError(
                        f"IPv4 subnet overlap in VRF '{vrf_name}': "
                        f"network '{net_name1}' ({subnet1_str}) overlaps with "
                        f"network '{net_name2}' ({subnet2_str}). "
                        f"Networks within the same VRF must not have overlapping subnets."
                    )

    # Check for IPv6 subnet overlaps within each VRF
    for vrf_name, subnets in vrf_subnets_v6.items():
        for i in range(len(subnets)):
            net_name1, subnet1_str = subnets[i]
            subnet1 = ipaddress.ip_network(subnet1_str, strict=False)

            for j in range(i + 1, len(subnets)):
                net_name2, subnet2_str = subnets[j]
                subnet2 = ipaddress.ip_network(subnet2_str, strict=False)

                if subnet1.overlaps(subnet2):
                    raise ValueError(
                        f"IPv6 subnet overlap in VRF '{vrf_name}': "
                        f"network '{net_name1}' ({subnet1_str}) overlaps with "
                        f"network '{net_name2}' ({subnet2_str}). "
                        f"Networks within the same VRF must not have overlapping subnets."
                    )

    # Count L2 vs L3 networks for validation summary
    l2_networks = [name for name, data in site_networks.items()
                   if not data.get("gateway") and not data.get("gateway6")]
    l3_networks = [name for name in network_name_list if name not in l2_networks]

    print(f"\nNetworks validation summary:")
    print(f"  - L2-only networks: {len(l2_networks)}")
    print(f"  - L3 routed networks: {len(l3_networks)}")
    if dhcpd_config:
        print(f"  - Networks with DHCP relay: {len(dhcpd_config)}")

    return site_networks, vrf_instances, network_name_list, dhcpd_config


# =============================================================================
# WAN / L3OUT PARSING - Continued in next part due to length
# =============================================================================

def _parse_l3out(wb, switches_config: List[Dict[str, Any]], vrf_instances: Dict[str, Any], site_networks: Dict[str, Any],
                 dhcpd_config: Dict[str, Any], auto_loopback_subnet: Optional[str], auto_loopback_subnet6: Optional[str],
                 core_as_border: bool) -> Dict[str, Any]:
    """
    Parse WAN sheet for external BGP peering configuration - STRICT validation

    Note: Row 1 contains comments, Row 2 contains headers, data starts at Row 3

    Returns: Dictionary mapping hostname to WAN configuration
    """
    # If WAN sheet doesn't exist, return empty dict
    if "WAN" not in wb.sheetnames:
        return {}

    sh = wb["WAN"]
    # Row 1 is comments, Row 2 is headers
    header = [str(c).strip().upper() if c else "" for c in next(sh.iter_rows(min_row=2, max_row=2, values_only=True))]

    required_columns = [
        "HOSTNAME", "VRF", "NETWORKNAME", "VLAN_ID", "INTERFACE", "TYPE", "IP", "PEER",
        "LOCAL_AS", "PEER_AS", "EXPORT_POLICY", "EXPORT_POLICY_DEFAULT_VRF_NETWORKS",
        "EXPORT_TERM_NAME", "EXPORT_TERM_PREFIX", "EXPORT_TERM_PROTOCOL", "EXPORT_TERM_ACTION",
        "IMPORT_POLICY", "IMPORT_TERM_NAME", "IMPORT_TERM_PREFIX", "IMPORT_TERM_PROTOCOL",
        "IMPORT_TERM_ACTION"
    ]

    for col in required_columns:
        if col not in header:
            raise Exception(f"WAN sheet missing required column: {col}")

    # Build lookup dictionaries
    switches_by_name = {sw["hostname"]: sw for sw in switches_config}
    l3out_config: Dict[str, Any] = {}

    # Determine which roles are allowed based on core_as_border
    if core_as_border:
        allowed_roles = {"core"}
        role_desc = "core"
    else:
        allowed_roles = {"border"}
        role_desc = "border"

    row_num = 3  # Data starts at row 3
    for row in sh.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            row_num += 1
            continue

        try:
            l3 = {header[i]: row[i] for i in range(min(len(header), len(row)))}

            # Validate hostname
            hostname = _validate_required(l3.get("HOSTNAME"), "HOSTNAME", f"WAN row {row_num}")
            hostname = str(hostname).strip()

            if hostname not in switches_by_name:
                raise ValueError(
                    f"HOSTNAME '{hostname}' in WAN row {row_num} not found in SWITCHES sheet"
                )

            sw_role = switches_by_name[hostname]["role"]
            if sw_role not in allowed_roles:
                raise ValueError(
                    f"HOSTNAME '{hostname}' in WAN row {row_num} has role '{sw_role}' but "
                    f"core_as_border={core_as_border} requires role to be '{role_desc}'. "
                    f"When core_as_border=true, only core switches can have WAN configuration. "
                    f"When core_as_border=false, only border switches can have WAN configuration."
                )

            # Parse and validate VRF
            vrf = _validate_required(l3.get("VRF"), "VRF", f"WAN row {row_num}")
            vrf = str(vrf).strip()

            # Validate VRF exists in NETWORKS sheet
            if vrf not in vrf_instances:
                available_vrfs = list(vrf_instances.keys()) if vrf_instances else []
                raise ValueError(
                    f"VRF '{vrf}' in WAN row {row_num} not found in NETWORKS sheet. "
                    f"WAN can only use VRFs that are defined in the NETWORKS sheet. "
                    f"Available VRFs: {', '.join(available_vrfs) if available_vrfs else 'none'}"
                )

            # Parse basic fields
            network_name = _validate_required(l3.get("NETWORKNAME"), "NETWORKNAME", f"WAN row {row_num}")
            network_name = str(network_name).strip()

            vlan_id = _i(l3.get("VLAN_ID"), "VLAN_ID", f"WAN row {row_num}")
            if vlan_id < 1 or vlan_id > 4094:
                raise ValueError(f"VLAN_ID {vlan_id} out of valid range (1-4094) in WAN row {row_num}")

            interface = _validate_required(l3.get("INTERFACE"), "INTERFACE", f"WAN row {row_num}")
            interface = str(interface).strip()

            port_type = _validate_required(l3.get("TYPE"), "TYPE", f"WAN row {row_num}")
            port_type = str(port_type).strip().lower()

            ip_with_cidr = _validate_required(l3.get("IP"), "IP", f"WAN row {row_num}")
            ip_with_cidr = str(ip_with_cidr).strip()

            # Validate and parse IP address with CIDR
            try:
                ip_iface = ipaddress.ip_interface(ip_with_cidr)
                ip_addr = str(ip_iface.ip)
                netmask = f"/{ip_iface.network.prefixlen}"
            except ValueError as e:
                raise ValueError(f"Invalid IP address '{ip_with_cidr}' in WAN row {row_num}: {e}")

            # Validate peer IP
            peer_ip = _validate_required(l3.get("PEER"), "PEER", f"WAN row {row_num}")
            peer_ip = str(peer_ip).strip()
            try:
                peer_addr = ipaddress.ip_address(peer_ip)
                peer_ip = str(peer_addr)
            except ValueError as e:
                raise ValueError(f"Invalid PEER IP '{peer_ip}' in WAN row {row_num}: {e}")

            # Parse BGP AS numbers
            local_as = _i(l3.get("LOCAL_AS"), "LOCAL_AS", f"WAN row {row_num}")
            peer_as = _i(l3.get("PEER_AS"), "PEER_AS", f"WAN row {row_num}")

            # Parse BFD_TIMER (optional)
            bfd_timer = None
            bfd_timer_raw = l3.get("BFD_TIMER")
            if bfd_timer_raw is not None and str(bfd_timer_raw).strip() != "":
                bfd_timer = _i(bfd_timer_raw, "BFD_TIMER", f"WAN row {row_num}")
                if bfd_timer < 1:
                    raise ValueError(f"BFD_TIMER must be >= 1 in WAN row {row_num}")

            # Parse export policy
            export_policy = _validate_required(l3.get("EXPORT_POLICY"), "EXPORT_POLICY", f"WAN row {row_num}")
            export_policy = str(export_policy).strip()

            # Parse EXPORT_POLICY_DEFAULT_VRF_NETWORKS flag
            export_default_vrf_networks = _b(
                l3.get("EXPORT_POLICY_DEFAULT_VRF_NETWORKS"),
                "EXPORT_POLICY_DEFAULT_VRF_NETWORKS",
                f"WAN row {row_num}"
            )

            # Parse export terms - only required if EXPORT_POLICY_DEFAULT_VRF_NETWORKS is FALSE
            export_term_names = []
            export_term_prefixes = []
            export_term_protocols = []
            export_term_actions = []

            if not export_default_vrf_networks:
                # Manual export terms are required when auto-generation is disabled
                export_term_names = _validate_required(l3.get("EXPORT_TERM_NAME"), "EXPORT_TERM_NAME", f"WAN row {row_num}")
                export_term_names = _clean_excel_string(str(export_term_names).strip()).split('\n')

                export_term_prefixes = _validate_required(l3.get("EXPORT_TERM_PREFIX"), "EXPORT_TERM_PREFIX", f"WAN row {row_num}")
                export_term_prefixes = _clean_excel_string(str(export_term_prefixes).strip()).split('\n')

                export_term_protocols = _validate_required(l3.get("EXPORT_TERM_PROTOCOL"), "EXPORT_TERM_PROTOCOL", f"WAN row {row_num}")
                export_term_protocols = _clean_excel_string(str(export_term_protocols).strip()).split('\n')

                export_term_actions = _validate_required(l3.get("EXPORT_TERM_ACTION"), "EXPORT_TERM_ACTION", f"WAN row {row_num}")
                export_term_actions = _clean_excel_string(str(export_term_actions).strip()).split('\n')

                # Validate export terms have same count
                if not (len(export_term_names) == len(export_term_prefixes) == len(export_term_protocols) == len(export_term_actions)):
                    raise ValueError(
                        f"WAN row {row_num}: Export policy terms count mismatch. "
                        f"EXPORT_TERM_NAME has {len(export_term_names)} entries, "
                        f"EXPORT_TERM_PREFIX has {len(export_term_prefixes)} entries, "
                        f"EXPORT_TERM_PROTOCOL has {len(export_term_protocols)} entries, "
                        f"EXPORT_TERM_ACTION has {len(export_term_actions)} entries. "
                        f"All must have the same number of newline-separated values."
                    )
            else:
                # When auto-generation is enabled, manual terms are optional but will be appended if provided
                export_term_name_raw = l3.get("EXPORT_TERM_NAME")
                if export_term_name_raw and str(export_term_name_raw).strip():
                    export_term_names = _clean_excel_string(str(export_term_name_raw).strip()).split('\n')

                    export_term_prefix_raw = l3.get("EXPORT_TERM_PREFIX")
                    if export_term_prefix_raw and str(export_term_prefix_raw).strip():
                        export_term_prefixes = _clean_excel_string(str(export_term_prefix_raw).strip()).split('\n')

                    export_term_protocol_raw = l3.get("EXPORT_TERM_PROTOCOL")
                    if export_term_protocol_raw and str(export_term_protocol_raw).strip():
                        export_term_protocols = _clean_excel_string(str(export_term_protocol_raw).strip()).split('\n')

                    export_term_action_raw = l3.get("EXPORT_TERM_ACTION")
                    if export_term_action_raw and str(export_term_action_raw).strip():
                        export_term_actions = _clean_excel_string(str(export_term_action_raw).strip()).split('\n')

                    # If any optional terms are provided, validate they all have the same count
                    if export_term_names or export_term_prefixes or export_term_protocols or export_term_actions:
                        if not (len(export_term_names) == len(export_term_prefixes) == len(export_term_protocols) == len(export_term_actions)):
                            raise ValueError(
                                f"WAN row {row_num}: Optional export policy terms count mismatch when EXPORT_POLICY_DEFAULT_VRF_NETWORKS=true. "
                                f"EXPORT_TERM_NAME has {len(export_term_names)} entries, "
                                f"EXPORT_TERM_PREFIX has {len(export_term_prefixes)} entries, "
                                f"EXPORT_TERM_PROTOCOL has {len(export_term_protocols)} entries, "
                                f"EXPORT_TERM_ACTION has {len(export_term_actions)} entries. "
                                f"If you provide manual terms alongside auto-generation, all must have the same number of values."
                            )

            # Parse import policy
            import_policy = _validate_required(l3.get("IMPORT_POLICY"), "IMPORT_POLICY", f"WAN row {row_num}")
            import_policy = str(import_policy).strip()

            import_term_names = _validate_required(l3.get("IMPORT_TERM_NAME"), "IMPORT_TERM_NAME", f"WAN row {row_num}")
            import_term_names = _clean_excel_string(str(import_term_names).strip()).split('\n')

            import_term_prefixes = _validate_required(l3.get("IMPORT_TERM_PREFIX"), "IMPORT_TERM_PREFIX", f"WAN row {row_num}")
            import_term_prefixes = _clean_excel_string(str(import_term_prefixes).strip()).split('\n')

            import_term_protocols = _validate_required(l3.get("IMPORT_TERM_PROTOCOL"), "IMPORT_TERM_PROTOCOL", f"WAN row {row_num}")
            import_term_protocols = _clean_excel_string(str(import_term_protocols).strip()).split('\n')

            import_term_actions = _validate_required(l3.get("IMPORT_TERM_ACTION"), "IMPORT_TERM_ACTION", f"WAN row {row_num}")
            import_term_actions = _clean_excel_string(str(import_term_actions).strip()).split('\n')

            # Validate import terms have same count
            if not (len(import_term_names) == len(import_term_prefixes) == len(import_term_protocols) == len(import_term_actions)):
                raise ValueError(
                    f"WAN row {row_num}: Import policy terms count mismatch. "
                    f"IMPORT_TERM_NAME has {len(import_term_names)} entries, "
                    f"IMPORT_TERM_PREFIX has {len(import_term_prefixes)} entries, "
                    f"IMPORT_TERM_PROTOCOL has {len(import_term_protocols)} entries, "
                    f"IMPORT_TERM_ACTION has {len(import_term_actions)} entries. "
                    f"All must have the same number of newline-separated values."
                )

            # Parse EXPORT_EXPLICT_DENY and IMPORT_EXPLICT_DENY flags (optional, default to True)
            export_explicit_deny = _b_optional(
                l3.get("EXPORT_EXPLICT_DENY"),
                "EXPORT_EXPLICT_DENY",
                f"WAN row {row_num}",
                default=True
            )

            import_explicit_deny = _b_optional(
                l3.get("IMPORT_EXPLICT_DENY"),
                "IMPORT_EXPLICT_DENY",
                f"WAN row {row_num}",
                default=True
            )

            # Validate all prefixes
            for i, prefix in enumerate(export_term_prefixes):
                prefix = prefix.strip()
                if not prefix:
                    raise ValueError(f"WAN row {row_num}: Empty prefix in EXPORT_TERM_PREFIX term {i+1}")
                try:
                    ipaddress.ip_network(prefix, strict=False)
                except ValueError as e:
                    raise ValueError(
                        f"WAN row {row_num}: Invalid prefix '{prefix}' in EXPORT_TERM_PREFIX term {i+1}: {e}"
                    )

            for i, prefix in enumerate(import_term_prefixes):
                prefix = prefix.strip()
                if not prefix:
                    raise ValueError(f"WAN row {row_num}: Empty prefix in IMPORT_TERM_PREFIX term {i+1}")
                try:
                    ipaddress.ip_network(prefix, strict=False)
                except ValueError as e:
                    raise ValueError(
                        f"WAN row {row_num}: Invalid prefix '{prefix}' in IMPORT_TERM_PREFIX term {i+1}: {e}"
                    )

            # Build export policy terms
            export_terms = []

            # Step 1: If EXPORT_POLICY_DEFAULT_VRF_NETWORKS is TRUE, generate default terms for all VRF networks
            if export_default_vrf_networks:
                # Get all networks in this VRF and build export terms for their subnets
                vrf_networks = vrf_instances.get(vrf, {}).get("networks", [])
                term_counter = 0

                for net_name in vrf_networks:
                    net_data = site_networks.get(net_name, {})

                    # Add IPv4 subnet term if present
                    if net_data.get("subnet"):
                        export_terms.append({
                            "matching": {
                                "prefix": [net_data["subnet"]]
                            },
                            "actions": {
                                "accept": True
                            },
                            "name": f"{vrf}_DF_{term_counter}"
                        })
                        term_counter += 1

                    # Add IPv6 subnet term if present
                    if net_data.get("subnet6"):
                        export_terms.append({
                            "matching": {
                                "prefix": [net_data["subnet6"]]
                            },
                            "actions": {
                                "accept": True
                            },
                            "name": f"{vrf}_DF_{term_counter}"
                        })
                        term_counter += 1

                # Step 2: Check if VRF has DHCP relay enabled - add loopback subnet terms
                vrf_has_dhcpv4 = False
                vrf_has_dhcpv6 = False

                for net_name in vrf_networks:
                    if net_name in dhcpd_config:
                        dhcp_entry = dhcpd_config[net_name]
                        # Check if this network has IPv4 DHCP relay
                        if dhcp_entry.get("servers"):  # IPv4 DHCP servers present
                            vrf_has_dhcpv4 = True
                        # Check if this network has IPv6 DHCP relay
                        if dhcp_entry.get("serversv6"):  # IPv6 DHCP servers present
                            vrf_has_dhcpv6 = True

                # Add DHCPv4 loopback subnet term if VRF has DHCPv4 and auto_loopback_subnet is defined
                if vrf_has_dhcpv4 and auto_loopback_subnet:
                    # Format as "10.200.0.0/24-32" meaning /24 subnet with /32 host routes
                    try:
                        loopback_net = ipaddress.ip_network(auto_loopback_subnet, strict=False)
                        loopback_prefix = f"{loopback_net.network_address}/{loopback_net.prefixlen}-32"

                        export_terms.append({
                            "matching": {
                                "prefix": [loopback_prefix]
                            },
                            "actions": {
                                "accept": True
                            },
                            "name": f"{vrf}_DF_DHCPv4"
                        })
                    except ValueError as e:
                        raise ValueError(
                            f"WAN row {row_num}: Invalid auto_loopback_subnet '{auto_loopback_subnet}' "
                            f"when generating DHCPv4 export term: {e}"
                        )

                # Add DHCPv6 loopback subnet term if VRF has DHCPv6 and auto_loopback_subnet6 is defined
                if vrf_has_dhcpv6 and auto_loopback_subnet6:
                    # Format as "2001:2000::/64-128" meaning /64 subnet with /128 host routes
                    try:
                        loopback_net6 = ipaddress.ip_network(auto_loopback_subnet6, strict=False)
                        loopback_prefix6 = f"{loopback_net6.network_address}/{loopback_net6.prefixlen}-128"

                        export_terms.append({
                            "matching": {
                                "prefix": [loopback_prefix6]
                            },
                            "actions": {
                                "accept": True
                            },
                            "name": f"{vrf}_DF_DHCPv6"
                        })
                    except ValueError as e:
                        raise ValueError(
                            f"WAN row {row_num}: Invalid auto_loopback_subnet6 '{auto_loopback_subnet6}' "
                            f"when generating DHCPv6 export term: {e}"
                        )

            # Step 3: Add user-defined export terms from Excel (these come after auto-generated terms)
            for i in range(len(export_term_names)):
                action = export_term_actions[i].strip().upper()
                # ACCEPT -> accept: true, REJECT -> accept: false
                accept_value = (action == "ACCEPT")

                export_terms.append({
                    "matching": {
                        "prefix": [export_term_prefixes[i].strip()],
                        "protocol": [export_term_protocols[i].strip().lower()]
                    },
                    "actions": {
                        "accept": accept_value
                    },
                    "name": export_term_names[i].strip()
                })

            # Step 4: Add explicit deny term at end of export policy if EXPORT_EXPLICT_DENY is True
            if export_explicit_deny:
                # Add <VRF>_DF_DENY_ALL term at the end
                export_terms.append({
                    "matching": {},
                    "actions": {
                        "accept": False
                    },
                    "name": f"{vrf}_DF_DENY_ALL"
                })

            # Build import policy terms
            import_terms = []
            for i in range(len(import_term_names)):
                action = import_term_actions[i].strip().upper()
                # ACCEPT -> accept: true, REJECT -> accept: false
                accept_value = (action == "ACCEPT")

                import_terms.append({
                    "matching": {
                        "prefix": [import_term_prefixes[i].strip()],
                        "protocol": [import_term_protocols[i].strip().lower()]
                    },
                    "actions": {
                        "accept": accept_value
                    },
                    "name": import_term_names[i].strip()
                })

            # Step 5: Add explicit deny term at end of import policy if IMPORT_EXPLICT_DENY is True
            if import_explicit_deny:
                # Add <VRF>_DF_DENY_ALL term at the end
                import_terms.append({
                    "matching": {},
                    "actions": {
                        "accept": False
                    },
                    "name": f"{vrf}_DF_DENY_ALL"
                })

            # Initialize hostname config if not exists
            if hostname not in l3out_config:
                l3out_config[hostname] = {
                    "networks": {},
                    "other_ip_configs": {},
                    "vrf_instances": {},
                    "port_config": {},
                    "bgp_config": {},
                    "routing_policies": {}
                }

            # Add network configuration
            l3out_config[hostname]["networks"][network_name] = {
                "vlan_id": str(vlan_id),
                "subnet": "",
                "subnet6": ""
            }

            # Add other_ip_configs
            l3out_config[hostname]["other_ip_configs"][network_name] = {
                "type": "static",
                "ip": ip_addr,
                "netmask": netmask
            }

            # Add VRF instance (copy complete network list from NETWORKS sheet and add WAN network)
            if vrf not in l3out_config[hostname]["vrf_instances"]:
                # Copy the complete VRF configuration from the NETWORKS sheet
                # This includes all networks that belong to this VRF, not just the WAN network
                original_vrf = vrf_instances.get(vrf, {})
                l3out_config[hostname]["vrf_instances"][vrf] = {
                    "networks": list(original_vrf.get("networks", [])),  # Copy all networks from NETWORKS sheet
                    "extra_routes": dict(original_vrf.get("extra_routes", {})),  # Copy IPv4 static routes
                    "extra_routes6": dict(original_vrf.get("extra_routes6", {}))  # Copy IPv6 static routes
                }

            # Add the WAN network if not already in the list
            if network_name not in l3out_config[hostname]["vrf_instances"][vrf]["networks"]:
                l3out_config[hostname]["vrf_instances"][vrf]["networks"].append(network_name)

            # Add port configuration - if interface already exists, append to networks list
            if interface not in l3out_config[hostname]["port_config"]:
                # First time seeing this interface - create new config
                l3out_config[hostname]["port_config"][interface] = {
                    "usage": port_type,
                    "critical": True,
                    "description": "",
                    "no_local_overwrite": True,
                    "speed": "auto",
                    "networks": [network_name],
                    "duplex": "auto",
                    "disable_autoneg": False
                }
            else:
                # Interface already exists - append network to the list if not already present
                if network_name not in l3out_config[hostname]["port_config"][interface]["networks"]:
                    l3out_config[hostname]["port_config"][interface]["networks"].append(network_name)

            # Add BGP configuration
            bgp_cfg = {
                "type": "external",
                "networks": [network_name],
                "local_as": local_as,
                "neighbors": {
                    peer_ip: {
                        "neighbor_as": peer_as
                    }
                },
                "export_policy": export_policy,
                "import_policy": import_policy
            }

            # Add BFD timer only if configured
            if bfd_timer is not None:
                bgp_cfg["bfd_minimum_interval"] = str(bfd_timer)

            l3out_config[hostname]["bgp_config"][vrf] = bgp_cfg

            # Add routing policies
            l3out_config[hostname]["routing_policies"][export_policy] = {
                "terms": export_terms
            }

            l3out_config[hostname]["routing_policies"][import_policy] = {
                "terms": import_terms
            }

        except Exception as e:
            raise Exception(f"Error parsing WAN row {row_num}: {e}")

        row_num += 1

    if l3out_config:
        print(f"\nWAN configuration parsed:")
        print(f"  - Switches with WAN: {len(l3out_config)}")
        print(f"  - Switches: {', '.join(l3out_config.keys())}")
    return l3out_config


def _build_pod_names(switches: List[Dict[str, Any]]) -> Dict[str, str]:
    """Extract unique pod numbers and create pod names"""
    pod_numbers = set()
    for sw in switches:
        if sw.get("pod") is not None:
            pod_numbers.add(sw["pod"])

    if not pod_numbers:
        return {}

    return {str(p): f"Pod {p}" for p in sorted(pod_numbers)}


def _validate_topology_consistency(switches_config: List[Dict[str, Any]], core_as_border: bool):
    """Validate that the topology is consistent with core_as_border setting"""
    border_switches = [sw for sw in switches_config if sw["role"] == "border"]
    core_switches = [sw for sw in switches_config if sw["role"] == "core"]

    if core_as_border and border_switches:
        raise Exception(
            f"Topology error: core_as_border=true but {len(border_switches)} border switch(es) defined. "
            f"When core_as_border=true, do not define separate border switches."
        )

    if not core_as_border and not border_switches:
        raise Exception(
            f"Topology error: core_as_border=false but no border switches defined. "
            f"When core_as_border=false, you must define switches with role='border'."
        )

    # When core acts as border, validate core switches do NOT have uplinks
    if core_as_border:
        for core_sw in core_switches:
            if core_sw["uplinks"]:
                raise Exception(
                    f"Topology error: core_as_border=true but core switch '{core_sw['hostname']}' has uplinks defined: {', '.join(core_sw['uplinks'])}. "
                    f"When core_as_border=true, core switches should NOT have uplinks (they are the border)."
                )
            if core_sw["uplink_ports"]:
                raise Exception(
                    f"Topology error: core_as_border=true but core switch '{core_sw['hostname']}' has UPLINK_PORTS defined: '{core_sw['uplink_ports']}'. "
                    f"When core_as_border=true, core switches should NOT have UPLINK_PORTS defined."
                )

    # When using dedicated borders, validate core switches have uplinks to borders
    if not core_as_border:
        border_hostnames = {sw["hostname"] for sw in border_switches}

        for core_sw in core_switches:
            if not core_sw["uplinks"]:
                raise Exception(
                    f"Topology error: Core switch '{core_sw['hostname']}' has no uplinks. "
                    f"When using dedicated borders (core_as_border=false), core switches must have uplinks to border switches."
                )

            # Verify uplinks point to border switches
            for uplink in core_sw["uplinks"]:
                if uplink not in border_hostnames:
                    raise Exception(
                        f"Topology error: Core switch '{core_sw['hostname']}' has uplink to '{uplink}', "
                        f"but '{uplink}' is not a border switch. Core uplinks must point to border switches."
                    )


def _validate_pod_structure(switches_config: List[Dict[str, Any]]):
    """
    Validate pod structure and detect topology type.
    Pods can be either:
    - 3-tier: Access -> Distribution -> Core
    - 2-tier: Access -> Core (no distribution layer)
    """
    # Group switches by pod
    pods: Dict[int, Dict[str, List[str]]] = {}

    for sw in switches_config:
        if sw.get("pod") is not None:
            pod_num = sw["pod"]
            if pod_num not in pods:
                pods[pod_num] = {"distribution": [], "access": []}

            if sw["role"] == "distribution":
                pods[pod_num]["distribution"].append(sw["hostname"])
            elif sw["role"] == "access":
                pods[pod_num]["access"].append(sw["hostname"])
            elif sw["role"] in ["core", "border"]:
                # Core and border switches should not have pods
                raise Exception(
                    f"Pod validation error: {sw['role'].capitalize()} switch '{sw['hostname']}' "
                    f"is assigned to pod {pod_num}. Core and border switches should not be assigned to pods."
                )

    # Validate each pod has required switches and detect topology
    for pod_num, roles in pods.items():
        # Each pod must have at least one access switch
        if not roles["access"]:
            raise Exception(
                f"Pod validation error: Pod {pod_num} has no access switches. "
                f"Each pod must have at least one access switch."
            )

        # Detect and log topology type
        if roles["distribution"]:
            # 3-tier topology: Access -> Distribution -> Core
            print(f"Pod {pod_num} structure validated (3-tier topology):")
            print(f"  - Distribution switches: {', '.join(roles['distribution'])}")
            print(f"  - Access switches: {', '.join(roles['access'])}")
        else:
            # 2-tier topology: Access -> Core (no distribution layer)
            print(f"Pod {pod_num} structure validated (2-tier topology):")
            print(f"  - Access switches: {', '.join(roles['access'])}")
            print(f"  - Note: No distribution switches - access switches connect directly to core")


def _find_existing_topology(mist: mistClient.Mist, site_id: str, topology_name: str) -> Optional[Dict[str, Any]]:
    """
    Search for an existing EVPN topology by name in the specified site
    Returns the topology dict if found, None otherwise
    """
    try:
        topos = mist.get(f"sites/{site_id}/evpn_topologies")
        if not topos:
            return None

        for topo in topos:
            if topo.get("name") == topology_name:
                return topo

        return None
    except Exception as e:
        print(f"Warning: Failed to retrieve existing topologies: {e}")
        return None


def _get_device_config(mist: mistClient.Mist, site_id: str, device_id: str) -> Dict[str, Any]:
    """
    Fetch current device configuration from Mist API
    Returns the complete device configuration dict
    """
    try:
        url = f"sites/{site_id}/devices/{device_id}"
        device = mist.get(url)
        return device if device else {}
    except Exception as e:
        print(f"Warning: Failed to retrieve device config for {device_id}: {e}")
        return {}


def _merge_port_configs(
    current_port_config: Dict[str, Any],
    new_evpn_ports: Dict[str, Any],
    new_wan_ports: Dict[str, Any],
    site_networks: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Intelligently merge port configurations, preserving user-configured ports.

    Strategy:
    1. Start with current port config (preserves everything)
    2. Filter out OLD WAN ports that are no longer in new WAN config
    3. Update/add EVPN ports (uplink/downlink)
    4. Update/add WAN ports

    This ensures that user-configured ports (like access ports, trunk ports, etc.)
    are never touched by the script.
    """
    merged_config = {}

    # Process each port from current config
    for port_name, port_cfg in current_port_config.items():
        usage = port_cfg.get("usage", "")
        networks_on_port = port_cfg.get("networks", [])

        # Check if this is a WAN port (usage=inet with networks)
        if usage == "inet" and networks_on_port:
            # Check if any networks on this port are WAN networks (NOT in site_networks)
            wan_networks_on_port = [net for net in networks_on_port if net not in site_networks]

            if wan_networks_on_port:
                # This port has WAN networks
                # Only preserve if this port is in the NEW WAN configuration
                if port_name in new_wan_ports:
                    # Port is still in new WAN config - will be updated below
                    pass
                else:
                    # Port is NOT in new WAN config - skip it (don't preserve old WAN port)
                    continue

        # Port is either:
        # - Not a WAN port (user-configured, EVPN, etc.) - preserve it
        # - A WAN port that's still in new WAN config - preserve it (will be updated below)
        merged_config[port_name] = port_cfg

    # Update with new EVPN ports (these take precedence)
    for port_name, port_cfg in new_evpn_ports.items():
        merged_config[port_name] = port_cfg

    # Update with new WAN ports (these also take precedence)
    for port_name, port_cfg in new_wan_ports.items():
        merged_config[port_name] = port_cfg

    return merged_config


def create_fabric(xlsx_path: str):
    print("Connecting to Mist API...")
    mist = mistClient.Mist(MIST_API_URL, MIST_TOKEN, MIST_ORGID)
    mh = mistHelpers.MistHelpers(mist)

    if not mist.test_connection():
        raise Exception("Failed to connect to Mist API")
    print("Successfully connected to API.")

    # Load Excel workbook
    try:
        wb = load_workbook(xlsx_path)
    except FileNotFoundError:
        raise Exception(f"Excel file not found: {xlsx_path}")
    except Exception as e:
        raise Exception(f"Error loading Excel file '{xlsx_path}': {e}")

    # Parse configuration with strict validation
    fabric = _parse_fabric(wb)
    switches_config = _parse_switches(wb)

    # Parse networks with IPv6 underlay flag for validation
    site_networks, vrf_instances, network_name_list, dhcpd_config = _parse_networks(
        wb,
        fabric["use_ipv6_underlay"]
    )

    # Parse WAN configuration (optional)
    l3out_config = _parse_l3out(
        wb,
        switches_config,
        vrf_instances,
        site_networks,
        dhcpd_config,
        fabric.get("auto_loopback_subnet"),
        fabric.get("auto_loopback_subnet6"),
        fabric["core_as_border"]
    )

    # Validate pod structure
    print("\n=== Validating pod structure ===")
    _validate_pod_structure(switches_config)

    # Validate topology consistency
    print("\n=== Validating topology consistency ===")
    _validate_topology_consistency(switches_config, fabric["core_as_border"])

    site_name = fabric["site_name"]
    topo_name = fabric["topology_name"]
    core_as_border = fabric["core_as_border"]

    # Lookup site ID
    site_id = None
    sites = mh.get_sites()
    for s in sites:
        if s.get("name") == site_name:
            site_id = s.get("id")
            break

    if not site_id:
        available_sites = [s.get("name") for s in sites if s.get("name")]
        raise Exception(f"Site '{site_name}' not found. Available sites: {', '.join(available_sites)}")

    print(f"Found site: {site_name} (ID: {site_id})")
    print(f"Topology mode: {'Core switches act as borders' if core_as_border else 'Dedicated border switches'}")
    print(f"IPv6 underlay: {'Enabled' if fabric['use_ipv6_underlay'] else 'Disabled'}")

    # =========================================================================
    # CHECK FOR EXISTING TOPOLOGY
    # =========================================================================
    existing_topology = _find_existing_topology(mist, site_id, topo_name)

    if existing_topology:
        print(f"\n*** EXISTING TOPOLOGY FOUND ***")
        print(f"Topology '{topo_name}' already exists with ID: {existing_topology['id']}")
        print(f"The fabric will be UPDATED according to the Excel configuration")
        mode = "UPDATE"
        topo_id = existing_topology["id"]
    else:
        print(f"\n*** NO EXISTING TOPOLOGY FOUND ***")
        print(f"Topology '{topo_name}' will be CREATED from scratch")
        mode = "CREATE"
        topo_id = None

    # Get devices in site
    devs = mh.get_switches(site_id)
    dev_by_name = {d.get("name"): d for d in devs}

    # Verify all switches exist
    missing_switches = []
    for sw in switches_config:
        if sw["hostname"] not in dev_by_name:
            missing_switches.append(sw["hostname"])

    if missing_switches:
        available_switches = list(dev_by_name.keys())
        raise Exception(
            f"Switch(es) not found in site '{site_name}': {', '.join(missing_switches)}\n"
            f"Available switches: {', '.join(available_switches)}"
        )

    # Build pod names
    pod_names = _build_pod_names(switches_config)

    # =========================================================================
    # STEP 1: Update site settings with networks, VRF instances, and DHCP config
    # =========================================================================
    print(f"\n=== STEP 1: {'Updating' if mode == 'UPDATE' else 'Creating'} site settings ===")
    site_setting_payload = {
        "vrf_instances": vrf_instances,
        "networks": site_networks,
        "port_usages": {}
    }

    # Add DHCP relay configuration if present
    if dhcpd_config:
        site_setting_payload["switch"] = {
            "dhcpd_config": dhcpd_config
        }
        print(f"DHCP relay configured for {len(dhcpd_config)} network(s)")

    print("Site settings payload:")
    print(f"  - VRF instances: {list(vrf_instances.keys())}")
    print(f"  - Networks: {len(site_networks)}")

    resp = mist.put(f"sites/{site_id}/setting", site_setting_payload)
    print("Site settings updated successfully")

    # =========================================================================
    # PRE-STEP 2: Fetch current device configurations BEFORE topology update
    # =========================================================================
    # CRITICAL: The EVPN topology update in STEP 2 will reset port configurations.
    # We must fetch and cache the current port configs NOW to preserve user-configured ports.
    print(f"\n=== PRE-STEP 2: Caching current device configurations ===")
    device_configs_cache = {}
    for sw_cfg in switches_config:
        dev = dev_by_name[sw_cfg["hostname"]]
        dev_id = dev.get("id")
        hostname = sw_cfg["hostname"]

        print(f"  Caching config for {hostname}...")
        current_config = _get_device_config(mist, site_id, dev_id)
        device_configs_cache[hostname] = current_config.get("port_config", {})

    print(f"Cached configurations for {len(device_configs_cache)} switches")

    # =========================================================================
    # STEP 2: Create or Update EVPN topology
    # =========================================================================
    print(f"\n=== STEP 2: {'Updating' if mode == 'UPDATE' else 'Creating'} EVPN topology ===")

    # Hardcoded overlay AS (Mist requirement)
    HARDCODED_OVERLAY_AS = 65000

    evpn_options = {
        "routed_at": "edge",
        "overlay": {"as": HARDCODED_OVERLAY_AS},
        "core_as_border": core_as_border,
        "underlay": {
            "as_base": fabric["bgp_as_pool_range_for_underlay"],
            "use_ipv6": fabric["use_ipv6_underlay"]
        }
    }

    # Add underlay subnet if defined (for both IPv4 and IPv6)
    if fabric.get("underlay_subnet"):
        evpn_options["underlay"]["subnet"] = fabric["underlay_subnet"]

    # Add optional subnets if present
    if fabric.get("auto_router_id_subnet"):
        evpn_options["auto_router_id_subnet"] = fabric["auto_router_id_subnet"]
    if fabric.get("auto_router_id_subnet6"):
        evpn_options["auto_router_id_subnet6"] = fabric["auto_router_id_subnet6"]
    if fabric.get("auto_loopback_subnet"):
        evpn_options["auto_loopback_subnet"] = fabric["auto_loopback_subnet"]
    if fabric.get("auto_loopback_subnet6"):
        evpn_options["auto_loopback_subnet6"] = fabric["auto_loopback_subnet6"]

    # Build switches array for topology
    switches = []
    for sw_cfg in switches_config:
        dev = dev_by_name[sw_cfg["hostname"]]
        mac = (dev.get("mac") or "").replace(":", "").replace("-", "").replace(".", "").lower()

        if not mac:
            raise Exception(f"Switch '{sw_cfg['hostname']}' has no MAC address")

        # Resolve uplink MACs
        uplink_macs = []
        for uplink_name in sw_cfg["uplinks"]:
            if uplink_name not in dev_by_name:
                raise Exception(f"Uplink switch '{uplink_name}' for '{sw_cfg['hostname']}' not found")
            uplink_mac = (dev_by_name[uplink_name].get("mac") or "").replace(":", "").replace("-", "").replace(".", "").lower()
            if not uplink_mac:
                raise Exception(f"Uplink switch '{uplink_name}' has no MAC address")
            uplink_macs.append(uplink_mac)

        # Resolve downlink MACs
        downlink_macs = []
        for downlink_name in sw_cfg["downlinks"]:
            if downlink_name not in dev_by_name:
                raise Exception(f"Downlink switch '{downlink_name}' for '{sw_cfg['hostname']}' not found")
            downlink_mac = (dev_by_name[downlink_name].get("mac") or "").replace(":", "").replace("-", "").replace(".", "").lower()
            if not downlink_mac:
                raise Exception(f"Downlink switch '{downlink_name}' has no MAC address")
            downlink_macs.append(downlink_mac)

        sw_entry = {
            "mac": mac,
            "role": sw_cfg["role"],
            "uplinks": uplink_macs,
            "config": {
                "port_config": {}
            }
        }

        # Add pod if specified (border switches should not have pods)
        if sw_cfg["pod"] is not None:
            sw_entry["pod"] = sw_cfg["pod"]

        # Add downlinks if present
        if downlink_macs:
            sw_entry["downlinks"] = downlink_macs

        # Configure ports
        if sw_cfg["uplink_ports"]:
            sw_entry["config"]["port_config"][sw_cfg["uplink_ports"]] = {"usage": "evpn_uplink"}

        if sw_cfg["downlink_ports"]:
            sw_entry["config"]["port_config"][sw_cfg["downlink_ports"]] = {"usage": "evpn_downlink"}

        switches.append(sw_entry)

    topo_payload = {
        "name": topo_name,
        "overwrite": True,
        "evpn_options": evpn_options,
        "switches": switches
    }

    # Only add pod_names if there are pods defined
    if pod_names:
        topo_payload["pod_names"] = pod_names

    if mode == "UPDATE":
        # Use PUT to update existing topology
        print(f"Updating existing topology '{topo_name}' (ID: {topo_id}) with {len(switches)} switches and {len(pod_names)} pod(s)")
        topo_resp = mist.put(f"sites/{site_id}/evpn_topologies/{topo_id}", topo_payload)
        print(f"EVPN topology updated successfully")
    else:
        # Use POST to create new topology
        print(f"Creating new topology '{topo_name}' with {len(switches)} switches and {len(pod_names)} pod(s)")
        topo_resp = mist.post(f"sites/{site_id}/evpn_topologies", topo_payload)
        topo_id = topo_resp.get("id")
        print(f"EVPN topology created with ID: {topo_id}")

    # =========================================================================
    # STEP 3: Configure individual devices
    # =========================================================================
    print(f"\n=== STEP 3: Configuring individual devices ===")

    # Categorize switches based on their role and topology mode
    if core_as_border:
        # Original behavior: access switches get anycast gateways, core switches might get WAN
        edge_switches = [sw for sw in switches_config if sw["role"] == "access"]
        non_edge_switches = [sw for sw in switches_config if sw["role"] != "access"]
    else:
        # New behavior: both access AND border switches get anycast gateways
        edge_switches = [sw for sw in switches_config if sw["role"] in ["access", "border"]]
        non_edge_switches = [sw for sw in switches_config if sw["role"] not in ["access", "border"]]

    # Configure edge switches (access or border) with anycast gateways
    for sw_cfg in edge_switches:
        dev = dev_by_name[sw_cfg["hostname"]]
        dev_id = dev.get("id")
        hostname = sw_cfg["hostname"]

        # Use CACHED device configuration (fetched BEFORE topology update) to preserve user-configured ports
        current_port_config = device_configs_cache.get(hostname, {})

        # Build other_ip_configs for anycast gateways
        other_ip_configs = {}
        for net_name, net_data in site_networks.items():
            entry = {}

            if net_data.get("gateway"):
                entry.update({
                    "type": "static",
                    "ip": net_data["gateway"],
                    "netmask": str(ipaddress.ip_network(net_data["subnet"], strict=False).netmask),
                    "evpn_anycast": True
                })

            if net_data.get("gateway6"):
                i6 = ipaddress.ip_interface(f"{net_data['gateway6']}/{net_data['subnet6'].split('/')[1]}")
                entry.update({
                    "type6": "static",
                    "ip6": net_data["gateway6"],
                    "netmask6": f"/{i6.network.prefixlen}"
                })

            if entry:
                other_ip_configs[net_name] = entry

        # Build new EVPN port configs
        evpn_port_config = {}
        if sw_cfg["uplink_ports"]:
            evpn_port_config[sw_cfg["uplink_ports"]] = {"usage": "evpn_uplink"}
        if sw_cfg["downlink_ports"]:
            evpn_port_config[sw_cfg["downlink_ports"]] = {"usage": "evpn_downlink"}

        # Check if this switch has WAN configuration
        # WAN is only applied to border switches when core_as_border=false
        # (border switches get both anycast gateways AND WAN when core_as_border=false)
        networks = {}
        bgp_config = {}
        routing_policies = {}
        vrf_config_dict = {}
        wan_port_config = {}

        if hostname in l3out_config and sw_cfg["role"] == "border":
            # Border switches can have WAN only when core_as_border=false
            l3out = l3out_config[hostname]

            # Merge WAN networks
            networks.update(l3out.get("networks", {}))

            # Merge WAN other_ip_configs (WAN IPs, not anycast)
            other_ip_configs.update(l3out.get("other_ip_configs", {}))

            # Get WAN port configs (will be merged later)
            wan_port_config = l3out.get("port_config", {})

            # Add BGP configuration
            bgp_config = l3out.get("bgp_config", {})

            # Add routing policies
            routing_policies = l3out.get("routing_policies", {})

            # Merge VRF instances from WAN
            vrf_config_dict = l3out.get("vrf_instances", {})

            print(f"  - WAN configuration detected for border switch: {hostname}")

        # Merge port configs intelligently (preserves user-configured ports)
        merged_port_config = _merge_port_configs(
            current_port_config,
            evpn_port_config,
            wan_port_config,
            site_networks
        )

        payload = {
            "other_ip_configs": other_ip_configs,
            "port_config": merged_port_config,
            "optic_port_config": sw_cfg["optic_config"],  # OPTIC CONFIG from SWITCHES
            "dhcpd_config": {"enabled": bool(dhcpd_config)},
            "vrf_config": {"enabled": bool(vrf_instances)}
        }

        # Add WAN-specific fields if present
        if networks:
            payload["networks"] = networks
        if bgp_config:
            payload["bgp_config"] = bgp_config
        if routing_policies:
            payload["routing_policies"] = routing_policies
        if vrf_config_dict:
            payload["vrf_instances"] = vrf_config_dict

        print(f"Configuring {sw_cfg['role']} switch: {sw_cfg['hostname']}")
        mist.put(f"sites/{site_id}/devices/{dev_id}", payload)

    # Configure core/distribution switches
    for sw_cfg in non_edge_switches:
        dev = dev_by_name[sw_cfg["hostname"]]
        dev_id = dev.get("id")
        hostname = sw_cfg["hostname"]

        # Use CACHED device configuration (fetched BEFORE topology update) to preserve user-configured ports
        current_port_config = device_configs_cache.get(hostname, {})

        # Build new EVPN port configs
        evpn_port_config = {}
        if sw_cfg["uplink_ports"]:
            evpn_port_config[sw_cfg["uplink_ports"]] = {"usage": "evpn_uplink"}
        if sw_cfg["downlink_ports"]:
            evpn_port_config[sw_cfg["downlink_ports"]] = {"usage": "evpn_downlink"}

        # Check if this switch has WAN configuration
        # WAN is only applied to core switches when core_as_border=true
        other_ip_configs = {}
        networks = {}
        bgp_config = {}
        routing_policies = {}
        vrf_config_dict = {}
        wan_port_config = {}

        if hostname in l3out_config and sw_cfg["role"] == "core":
            # Core switches can have WAN only when core_as_border=true
            l3out = l3out_config[hostname]

            # Add WAN networks
            networks.update(l3out.get("networks", {}))

            # Add WAN other_ip_configs
            other_ip_configs.update(l3out.get("other_ip_configs", {}))

            # Get WAN port configs (will be merged later)
            wan_port_config = l3out.get("port_config", {})

            # Add BGP configuration
            bgp_config = l3out.get("bgp_config", {})

            # Add routing policies
            routing_policies = l3out.get("routing_policies", {})

            # Merge VRF instances from WAN
            vrf_config_dict = l3out.get("vrf_instances", {})

            print(f"  - WAN configuration detected for core switch: {hostname}")

        # Merge port configs intelligently (preserves user-configured ports)
        merged_port_config = _merge_port_configs(
            current_port_config,
            evpn_port_config,
            wan_port_config,
            site_networks
        )

        payload = {
            "other_ip_configs": other_ip_configs,
            "port_config": merged_port_config,
            "optic_port_config": sw_cfg["optic_config"],  # OPTIC CONFIG from SWITCHES
        }

        # Enable vrf_config if switch has WAN configuration (core with core_as_border=true)
        # Otherwise disable it for non-edge switches
        if hostname in l3out_config:
            payload["vrf_config"] = {"enabled": True}
        else:
            payload["-vrf_config"] = True

        # Add WAN-specific fields if present
        if networks:
            payload["networks"] = networks
        if bgp_config:
            payload["bgp_config"] = bgp_config
        if routing_policies:
            payload["routing_policies"] = routing_policies
        if vrf_config_dict:
            payload["vrf_instances"] = vrf_config_dict

        print(f"Configuring {sw_cfg['role']} switch: {sw_cfg['hostname']}")
        mist.put(f"sites/{site_id}/devices/{dev_id}", payload)

    print(f"\n=== Fabric {'update' if mode == 'UPDATE' else 'creation'} complete ===")
    print(f"Mode: {mode}")
    print(f"Topology: {topo_name}")
    print(f"Topology ID: {topo_id}")
    print(f"Site: {site_name}")
    print(f"Mode: {'Core as Border' if core_as_border else 'Dedicated Border'}")
    print(f"Switches configured: {len(switches_config)}")
    if not core_as_border:
        border_count = len([sw for sw in switches_config if sw["role"] == "border"])
        print(f"  - Border switches: {border_count}")
    print(f"Networks: {len(site_networks)}")
    print(f"VRFs: {len(vrf_instances)}")
    print(f"Pods: {len(pod_names)}")
    if dhcpd_config:
        print(f"DHCP relay enabled on edge switches: {len(dhcpd_config)} network(s) configured")
    if l3out_config:
        print(f"WAN external BGP peering configured on {len(l3out_config)} switch(es): {', '.join(l3out_config.keys())}")


def main():
    if len(sys.argv) >= 2:
        xlsx = sys.argv[1]
    else:
        xlsx = spreadsheetname

    try:
        create_fabric(xlsx)
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
